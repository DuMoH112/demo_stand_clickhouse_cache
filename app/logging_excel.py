import os
from openpyxl import (
    Workbook,
    load_workbook
)


class LoggingResultDBToExcel():
    def __init__(self, path):
        self.files = []
        self.path = path
        path = self.path.split('/')
        self.dir, self.filename = "/".join(path[:-1]), path[-1]

        if self.__check_file():
            self.__load_file()
        else:
            self.__init_new_file()

    def __enter__(self):
        return self

    def __exit__(self, exc_type, exc_value, traceback):
        self.save()
        for file in self.files:
            os.unlink(file)

    def __init_new_file(self):
        self.wb = Workbook()
        self.main_table = self.wb.active
        self.main_table.title = 'Main'
        self.postgres_table = self.wb.create_sheet(title="Postgres")
        self.create_table_for_data(self.postgres_table)
        self.clickhouse_table = self.wb.create_sheet(title="Clickhouse")
        self.create_table_for_data(self.clickhouse_table)

    def __load_file(self):
        self.wb = load_workbook(self.path)
        self.postgres_table = self.wb['Postgres']
        self.clickhouse_table = self.wb['Clickhouse']

    def __check_file(self):
        for file in os.listdir(path=self.dir):
            if file == self.filename:
                return True
        
        return False

    def save(self):
        self.wb.save(filename = self.path)

    def close(self):
        self.save()
    
    def create_table_for_data(self, wt):
        wt.cell(column=1, row=1, value="Count data in tables")
        wt.cell(column=2, row=1, value="select_sum_value_in_all_files")
        wt.cell(column=3, row=1, value="select_min_and_max_dt")
        wt.cell(column=4, row=1, value="select_sum_value_in_one_files")
        wt.cell(column=5, row=1, value="select_without_filters")
        for i in range(1, 11):
            wt.cell(column=5 + i, row=1, value=f"select_with_{i}_filters")

    def get_last_column(self, table):
        return len([i for i in table])
